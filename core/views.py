from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Alumni
from .serializers import AlumniSerializer
from .permissions import IsAdminOrTechnicalHead
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO


class AlumniViewSet(viewsets.ModelViewSet):
    """
    Alumni management viewset
    """
    queryset = Alumni.objects.all()
    serializer_class = AlumniSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    
    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            permission_classes = [IsAdminOrTechnicalHead]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        """
        Filter alumni based on query parameters
        """
        queryset = Alumni.objects.all()
        
        # Filter by graduation year
        year = self.request.query_params.get('year')
        if year:
            queryset = queryset.filter(year_of_passout=year)
        
        # Filter by company
        company = self.request.query_params.get('company')
        if company:
            queryset = queryset.filter(current_workplace__icontains=company)
        
        # Filter by location
        location = self.request.query_params.get('location')
        if location:
            queryset = queryset.filter(current_location__icontains=location)
        
        # Filter by mentorship availability
        mentorship = self.request.query_params.get('mentorship')
        if mentorship:
            queryset = queryset.filter(willing_to_mentor=mentorship.lower() == 'true')
        
        return queryset.order_by('-year_of_passout', 'last_name')
    
    @action(detail=False, methods=['get'])
    def featured(self, request):
        """Get featured alumni"""
        featured_alumni = self.get_queryset().filter(willing_to_mentor=True)[:6]
        serializer = self.get_serializer(featured_alumni, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get alumni statistics"""
        queryset = self.get_queryset()
        
        # Group by graduation year
        year_stats = {}
        for alumni in queryset:
            year = alumni.year_of_passout
            if year not in year_stats:
                year_stats[year] = 0
            year_stats[year] += 1
        
        # Group by company
        company_stats = {}
        for alumni in queryset.filter(current_workplace__isnull=False):
            company = alumni.current_workplace
            if company not in company_stats:
                company_stats[company] = 0
            company_stats[company] += 1
        
        # Sort companies by count
        top_companies = sorted(company_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        
        return Response({
            'total_alumni': queryset.count(),
            'mentors_available': queryset.filter(willing_to_mentor=True).count(),
            'year_wise_distribution': year_stats,
            'top_companies': dict(top_companies)
        })
    
    @action(detail=False, methods=['get'], permission_classes=[IsAdminOrTechnicalHead])
    def export_excel(self, request):
        """Export alumni data to Excel"""
        queryset = self.get_queryset()
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alumni Data"
        
        # Define headers
        headers = [
            'Name', 'Email', 'Mobile', 'Branch', 'Admission Year', 'Passout Year',
            'CGPA', 'Current Workplace', 'Job Title', 'Location', 'LinkedIn',
            'Willing to Mentor', 'Achievements'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, alumni in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=alumni.full_name)
            ws.cell(row=row, column=2, value=alumni.email)
            ws.cell(row=row, column=3, value=alumni.mobile_number)
            ws.cell(row=row, column=4, value=alumni.branch)
            ws.cell(row=row, column=5, value=alumni.year_of_admission)
            ws.cell(row=row, column=6, value=alumni.year_of_passout)
            ws.cell(row=row, column=7, value=str(alumni.cgpa) if alumni.cgpa else '')
            ws.cell(row=row, column=8, value=alumni.current_workplace or '')
            ws.cell(row=row, column=9, value=alumni.job_title or '')
            ws.cell(row=row, column=10, value=alumni.current_location or '')
            ws.cell(row=row, column=11, value=alumni.linkedin_url or '')
            ws.cell(row=row, column=12, value='Yes' if alumni.willing_to_mentor else 'No')
            ws.cell(row=row, column=13, value=alumni.achievements or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="alumni_data.xlsx"'
        return response
